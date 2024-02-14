import openpyxl

# Read data in aCell of Excel File

def read_data_of_cell(path,row,column):

    worksheet=openpyxl.load_workbook(path)
    sheet=worksheet.active
    cell_data=sheet.cell(row=row,column=column)
    return cell_data.value

worksheet=openpyxl.load_workbook('C:\\Users\\vipul\\PycharmProjects\\Swag_Labs\\src\\resources\\Data\\Valid Credentials.xlsx')
sheet=worksheet.active
cell_data=sheet.cell(row=2,column=2)
print(cell_data.value)
